#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time   : 2024/9/6 10:51
# @Author : Carey
# @File : StorageUtils.py
# @Description
import platform
import openpyxl
import os
from openpyxl import Workbook


class StorageUtils():


    def __init__( self ):
        system = platform.system()
        if 'Windows' == system:
            self.folder = os.path.expanduser( '~/Document/Download' )
        if system in ['Darwin','Linux']:
            self.folder = os.path.expanduser( '~' )

        self.shopname = None


    def initFile(self, data):

        """
        初始化
        """
        self.ptId = data[ 'ptid' ]
        self.ptCode = data[ 'code' ]
        self.memId =  data[ 'memid' ]
        if 'shopname' in data:
            self.shopname = data[ 'shopname' ]

        if '1688' == self.ptCode:
            self.fname = f"阿里巴巴【{self.memId}】-店铺商品数据.xlsx"
            if None != self.shopname and len(self.shopname) > 0:
                self.fname = f"阿里巴巴_{self.shopname}【{self.memId}】-店铺商品数据.xlsx"
        if 'Vvic' == self.ptCode:
            self.fname = f"搜款网【{self.memId}】-店铺商品数据.xlsx"
            if None != self.shopname and len(self.shopname) > 0:
                self.fname = f"搜款网_{self.shopname}【{self.memId}】-店铺商品数据.xlsx"
        if 'taobao' == self.ptCode:
            self.fname = f"淘宝【{self.memId}】-店铺商品数据.xlsx"
            if None != self.shopname and len(self.shopname) > 0:
                self.fname = f"淘宝_{self.shopname}【{self.memId}】-店铺商品数据.xlsx"
        if 'tmall' == self.ptCode:
            self.fname = f"天猫【{self.memId}】-店铺商品数据.xlsx"
            if None != self.shopname and len(self.shopname) > 0:
                self.fname = f"天猫_{self.shopname}【{self.memId}】-店铺商品数据.xlsx"

        return True

    """
    写入方式基类
    """
    def writToExcel( self, driver, filePath, datas ):
        """
        写入 Excel - 文件
        """
        if False == os.path.exists( filePath ):
            wb = Workbook()
            ws = wb.active

            ws['A1'] = '编号'
            ws['B1'] = '商品名称'
            ws['C1'] = '售价'
            ws['D1'] = '上架日期'
            ws['E1'] = '收藏数'
            ws['F1'] = '销量'
            ws['G1'] = '链接'
        else:
            wb = openpyxl.load_workbook( filePath )
            ws = wb.active

        for i in datas:
            item = driver.item( i )
            if not item['values'] or len( item['values'] ) <=0 :
                continue

            if 1 == self.ptId:
                url = f"https://detail.1688.com/offer/{item['values'][0]}.html"
            elif 2 == self.ptId:
                url = f"https://item.taobao.com/item.htm?id={item['values'][0]}"
            elif 3 == self.ptId:
                url = f"https://item.taobao.com/item.htm?id={item['values'][0]}"
            elif 4 == self.ptId:
                url = f"https://detail.tmall.com/item.htm?id={item['values'][0]}"
            else:
                continue

            sign = item['values'][0]
            if item[ 'text' ] and len( item[ 'text' ] ) > 0:
                sign = item[ 'text' ]

            write = [sign, item['values'][1], item['values'][2], item['values'][3], item['values'][4], item['values'][5], url]
            ws.append(write)

        wb.save( filePath )
        return True


storageUtils = StorageUtils()
